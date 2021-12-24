#!/usr/bin/env python3
# coding=utf-8

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import colors

info = ['¥64.14', '-0.36  -0.56%', '\xa08.52 万球友关注', '交易中', ' 12-23 14:54:30 北京时间', '最高：', '64.98', '今开：', '64.01', '涨停：', '77.40', '成交量：', '54662手', '最低：', '63.46', '昨收：', '64.50', '跌停：', '51.60', '成交额：', '3.51亿', '换手：', '2.10%', '盘后量：', '--', '量比：', '0.65', '总市值：', '212.30亿', '振幅：', '2.36%', '盘后额：', '--', '委比：', '61.49%', '流通值：', '167.06亿', '市盈率(动)：', '40.92', '市盈率(TTM)：', '50.80', '每股收益：', '1.26', '股息(TTM)：', '0.40', '市盈率(静)：', '103.69', '市净率：', '8.07', '每股净资产：', '7.95', '股息率(TTM)：', '0.62%', '总股本：', '3.31亿', '52周最高：', '116.61', '质押率：', '--', '盈利情况：', '已盈利', '流通股：', '2.60亿', '52周最低：', '26.17', '商誉/净资产：', '0.02%', '注册制：', '否', '投票权：', '无差异', 'VIE结构：', '否', '货币单位：', 'CNY']

index = []
for i in zip(info[3::2],info[4::2]):
    index.append(list(i))

# 创建 Excel 文件对象
wb = Workbook()
ws1 = wb.active

#定制一个字体样式对象
ft = Font(bold=True, size=16, color='1E90FF')

ws1.title = '股票'

# 更改工作表标签的背景色,值是RRGGBB颜色代码
# http://www.sioe.cn/yingyong/yanse-rgb-16/
#ws1.sheet_properties.tabColor = "0000FF"

# 设置标题的内容和字体样式
ws1.cell(row=1,column=1,value="指数").font = ft
ws1.cell(row=1,column=2,value="值").font = ft

ws1.cell(row=2,column=1,value="全志科技").font = Font(size=14)
ws1.cell(row=2,column=2,value=info[0]).font = Font(size=14)
#ws1.cell(row=1,column=3,value="统计时间").font = ft

# 设置列宽
ws1.column_dimensions['A'].width = 16
ws1.column_dimensions['B'].width = 32
#ws1.column_dimensions['C'].width = 22

# 获取到所有的行以及每行的所有列
rows = ws1.iter_rows(min_row=3, max_col=2, max_row=len(index))

#for items in index:
for row, items in zip(rows, tuple(index)):
    #print(row, items)
    for cell, item in zip(row, items):
        cell.value = item
        cell.font = Font(size=14)
#        print(cell, item)

# 保存文件到硬盘
wb.save('qzkj.xlsx')
