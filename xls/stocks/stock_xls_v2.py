#!/usr/bin/env python3
# coding=utf-8

import time
import requests
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import colors

class Spider(object):

    def loadPage(self, code):
        """下载页面"""
        # 获取首页
        if int(code) < 600000:
            url = 'https://xueqiu.com/S/SZ' + code
        else:
            url = 'https://xueqiu.com/S/SH' + code

        # 财务报表页面
        cwbb_url = 'https://q.stock.sohu.com/cn/%s/cwzb.shtml' % str(code)
        headers = {"User-Agent":"Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (sKHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36"}

        # 解析主页信息
        res = requests.get(url, headers=headers)
        html = etree.HTML(res.text)
        stime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        name = html.xpath('//*[@id="app"]/div[2]/div[2]/div[1]//text()')    # 用于获取股票名字
        price = html.xpath('//*[@id="app"]/div[2]/div[2]/div/div/div/div/strong/text()')    # 用于获取股票价格
        info = html.xpath('//*[@id="app"]/div[2]/div[2]/div/table//text()')    # 用于获取股票基本信息

        # 解析财务报表页
        cwbb_res = requests.get(cwbb_url, headers=headers)
        cwbb_html = etree.HTML(cwbb_res.text)
        cwbb_zj = cwbb_html.xpath('/html/body/div[4]/div[2]/div[2]/div[2]/div/div[2]/table//tr[31]/td[1]/text()')    #总计
        growrate = cwbb_html.xpath('/html/body/div[4]/div[2]/div[2]/div[2]/div/div[2]/table//tr[25]/td[2]/span/text()')    # 净利润增长率
        fzpers = cwbb_html.xpath('/html/body/div[4]/div[2]/div[2]/div[2]/div/div[2]/table//tr[11]/td[1]/text()')    # 资产负债率
        fzper = ''.join(fzpers).strip()
        pers = []
        pers.append("%.2f%%" % float(fzper))    # 加上%
        pers.append(stime)
      
        if 600000 > int(code) > 300000: 
            stock = ( name + price + info[23:24] + info[31:32] + info[41:42] + info[33:34] + info[43:44] + growrate + pers)
        else:
            stock = ( name + price + info[23:24] + info[31:32] + info[41:42] + info[33:34] + info[43:44] + growrate + pers)
        
        return stock

    def writePage(self, stock):

        # 创建 Excel 文件对象
        wb = Workbook()
        ws1 = wb.active

        # 定制一个字体样式对象
        ft = Font(bold=True, size=16, color='1E90FF')

        # 设置标题的内容和字体样式
        ws1.title = '股票基本面'
        ws1.cell(row=1,column=1,value="名称").font = ft
        ws1.cell(row=1,column=2,value="价格").font = ft
        ws1.cell(row=1,column=3,value="市值").font = ft
        ws1.cell(row=1,column=4,value="流通值").font = ft
        ws1.cell(row=1,column=5,value="市盈率(静)").font = ft
        ws1.cell(row=1,column=6,value="市盈率(动)").font = ft
        ws1.cell(row=1,column=7,value="市净率").font = ft
        ws1.cell(row=1,column=8,value="净利润增长率").font = ft
        ws1.cell(row=1,column=9,value="资产负债率").font = ft
        ws1.cell(row=1,column=10,value="时间").font = ft

        # 设置列宽
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 16
        ws1.column_dimensions['C'].width = 16
        ws1.column_dimensions['D'].width = 18
        ws1.column_dimensions['E'].width = 20
        ws1.column_dimensions['F'].width = 20
        ws1.column_dimensions['G'].width = 20
        ws1.column_dimensions['H'].width = 20
        ws1.column_dimensions['I'].width = 20
        ws1.column_dimensions['J'].width = 28

        # 写入表格
        for col, sk in zip(range(1,11), stock):
            ws1.cell(row=2,column=col,value=sk).font = Font(size=14)

        # 保存文件到硬盘
        wb.save(stock[0] + '.xlsx')

if __name__=="__main__":
    #code = input("请输入股票代码：")
    for code in ['300458','002641']:
        spider = Spider()
        stock = spider.loadPage(code)
        spider.writePage(stock)
