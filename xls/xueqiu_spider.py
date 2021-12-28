#!/usr/bin/env python3
# coding=utf-8

import requests
import re,os
from lxml import etree

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import colors

class Spider(object):

    def loadPage(self, code):
        """下载页面"""
        #首页
        url = 'https://xueqiu.com/S/' + code
        print(url)
        headers = {"User-Agent":"Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (sKHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36"}
        res = requests.get(url, headers=headers)
        html = etree.HTML(res.text)
        stock = html.xpath('//*[@id="app"]/div[2]/div[2]//text()')
        info = html.xpath('//*[@id="app"]/div[2]/div[2]/div[4]//text()')
        
        self.writePage(stock, info, code)

    def writePage(self, stock, info, code):

        index = []
        for i in zip(info[3::2],info[4::2]):
            index.append(list(i))

		# 创建 Excel 文件对象
        wb = Workbook()
        ws1 = wb.active

		#定制一个字体样式对象
        ft = Font(bold=True, size=16, color='1E90FF')

		# 设置标题的内容和字体样式
        ws1.title = code
        ws1.cell(row=1,column=1,value="指数").font = ft
        ws1.cell(row=1,column=2,value="数据").font = ft

        ws1.cell(row=2,column=1,value=stock[0]).font = Font(size=14)
        ws1.cell(row=2,column=2,value=info[0]).font = Font(size=14)

		# 设置列宽
        ws1.column_dimensions['A'].width = 16
        ws1.column_dimensions['B'].width = 32

		# 获取到所有的行以及每行的所有列
        rows = ws1.iter_rows(min_row=3, max_col=2, max_row=len(index))

        for row, items in zip(rows, tuple(index)):
            for cell, item in zip(row, items):
                cell.value = item
                cell.font = Font(size=14)

		# 保存文件到硬盘
        wb.save(stock[0] + '.xlsx')

if __name__=="__main__":
    code = input("请输入股票完整代码：")
    spider = Spider()
    spider.loadPage(code)
